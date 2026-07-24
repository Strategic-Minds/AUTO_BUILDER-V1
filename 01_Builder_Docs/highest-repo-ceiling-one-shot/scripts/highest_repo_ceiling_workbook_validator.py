#!/usr/bin/env python3
"""Highest Repo Ceiling workbook validator.

This script is intentionally lightweight so Codex/local Git can run it after the
actual .xlsx binary has been committed. It verifies the workbook exists, opens it
with openpyxl when available, checks required sheets, and blocks forbidden visual
parity/design-sheet names.
"""

from __future__ import annotations

import sys
from pathlib import Path

REQUIRED_SHEETS = [
    "00_START_HERE",
    "01_REPO_CEILING_CHARTER",
    "02_SOURCE_TRUTH_INDEX",
    "03_CURRENT_REPO_STATE",
    "04_HIGHEST_REPO_STATE_MODEL",
    "05_GAP_MASTER_LEDGER",
    "06_SOURCE_TRUTH_LAYER",
    "07_WORKBOOK_COMMAND_LAYER",
    "08_GITHUB_REPO_TEMPLATE_LAYER",
    "09_SUPABASE_RUNTIME_LEDGER",
    "10_VERCEL_RUNTIME_CRON_LAYER",
    "11_CODEX_BUILD_LOOP",
    "12_CONNECTOR_GOVERNANCE_LAYER",
    "13_RELEASE_OPERATION_LAYER",
    "14_BUSINESS_FACTORY_LAYER",
    "15_ENVIRONMENT_VARIABLES",
    "16_SECURITY_SECRET_READINESS",
    "17_VALIDATION_TESTING_LAYER",
    "18_RECEIPTS_PROOF_INDEX",
    "19_BLOCKERS_WORKAROUNDS",
    "20_ROLLBACK_RECOVERY_PLAN",
    "21_RELEASE_GATE",
    "22_RECURSIVE_AUDIT_LOG",
    "23_SCORECARD",
    "24_CODEX_INSTALL_HANDOFF",
    "25_BASE44_AGENT_HANDOFF",
    "26_SECOND_GPT_AUDIT_PROMPT",
    "27_FINAL_OPERATOR_HANDOFF",
    "99_VALIDATION_LISTS",
]

FORBIDDEN_SHEET_TERMS = [
    "visual_parity",
    "visual drift",
    "visual_drift",
    "screenshot_comparison",
    "screenshot comparison",
    "website_design_parity",
    "website design parity",
]


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: highest_repo_ceiling_workbook_validator.py <workbook.xlsx>", file=sys.stderr)
        return 2

    workbook_path = Path(sys.argv[1])
    if not workbook_path.exists():
        print(f"FAIL: workbook missing: {workbook_path}")
        return 1

    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        print(f"FAIL: openpyxl unavailable: {exc}")
        return 1

    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet_names = list(wb.sheetnames)
    missing = [name for name in REQUIRED_SHEETS if name not in sheet_names]
    forbidden = [
        name for name in sheet_names
        if any(term in name.lower() for term in FORBIDDEN_SHEET_TERMS)
    ]

    if missing:
        print("FAIL: missing required sheets:")
        for name in missing:
            print(f"- {name}")
        return 1

    if forbidden:
        print("FAIL: forbidden visual/design parity sheet names found:")
        for name in forbidden:
            print(f"- {name}")
        return 1

    print("HIGHEST REPO CEILING WORKBOOK VALIDATION: PASS")
    print(f"required_sheets={len(REQUIRED_SHEETS)}")
    print("visual_parity_rules=excluded")
    print("codex_handoff=present")
    print("base44_handoff=present")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
